from openpyxl import load_workbook
from pymongo import MongoClient
from config import sheet
from datetime_converter import to_datetime
from config import mongodb


load_wb = load_workbook('./config/부산대 시간표.xlsx', data_only=True)


def _get_schedules(ws, section, start_row, end_row, direction_info):
    schedules = []
    for i in range(start_row, end_row + 1):
        line_name = ws.cell(row=i, column=direction_info.LINE_NAME).value
        line_number = ws.cell(row=i, column=direction_info.LINE_NUMBER).value
        if (line_number is not None) and (line_number != '역'):
            line_name = f'{line_number.replace("번", "")}{line_name}'

        json = {
            'line': {
                'name': line_name,
            },
            'depart_time': to_datetime(ws.cell(row=i, column=direction_info.DEPART_TIME).value),
            'arrive_time': to_datetime(ws.cell(row=i, column=direction_info.ARRIVE_TIME).value),
            'direction': direction_info.DIRECTION,
            'section': section
        }
        schedules.append(json)
    return schedules


def get_schedules(sheet_info):
    ws = load_wb[sheet_info.SHEET_NAME]
    schedules = []
    schedules += _get_schedules(ws, sheet_info.SECTION, sheet_info.START_ROW, sheet_info.END_ROW, sheet.ToCampus)
    schedules += _get_schedules(ws, sheet_info.SECTION, sheet_info.START_ROW, sheet_info.END_ROW, sheet.ToStation)
    return schedules


total = []
total += get_schedules(sheet.Weekday)
total += get_schedules(sheet.Holiday)
total += get_schedules(sheet.CampusOnly)


client = MongoClient(host=mongodb.host, username=mongodb.username, password=mongodb.password)
db = client[mongodb.db]

db.schedule.insert_many(total)

